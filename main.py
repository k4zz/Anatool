import sys
import getopt
import openpyxl
import logging
import queue
import tkinter as tk
from tkinter import ttk, VERTICAL, HORIZONTAL, N, S, E, W, LEFT, RIGHT, Label, Button, Entry, IntVar, StringVar
from tkinter import filedialog as fd
from tkinter.scrolledtext import ScrolledText

# Logger
logger = logging.getLogger(__name__)


class QueueHandler(logging.Handler):
    def __init__(self, log_queue):
        super().__init__()
        self.log_queue = log_queue

    def emit(self, record):
        self.log_queue.put(record)


# Data containers
class Protocol:
    def __init__(self, number, names, row):
        self.number = number
        self.names = names
        self.row = row

    def __repr__(self):
        return f"Protokol(number={self.number}, names={self.names})"

    def add_names(self, name):
        self.names.extend(name)


class Collation:
    def __init__(self, name, numbers, row):
        self.name = name
        self.numbers = numbers
        self.row = row

    def __repr__(self):
        return f"Zestawienie(name='{self.name}', numbers='{self.numbers})"

    def add_positions(self, positions):
        self.numbers.extend(positions)


class Settings:
    def __init__(self):
        self.__protocol_number_column_data: str = "A"
        self.__protocol_names_column_data: str = "C"
        self.__collation_name_column_data: str = "B"
        self.__collation_numbers_column_data: str = "C"

    def __convert_to_iterator(self, string: str):
        return ord(string.lower()) - 97

    def update_settings(self, pro_num_col: str, pro_name_col: str, coll_name_col: str, coll_num_col: str):
        self.__protocol_number_column_data = pro_num_col
        self.__protocol_names_column_data = pro_name_col
        self.__collation_name_column_data = coll_name_col
        self.__collation_numbers_column_data = coll_num_col

    def protocol_number_column(self):
        return self.__convert_to_iterator(self.__protocol_number_column_data)

    def protocol_names_column(self):
        return self.__convert_to_iterator(self.__protocol_names_column_data)

    def collation_name_column(self):
        return self.__convert_to_iterator(self.__collation_name_column_data)

    def collation_numbers_column(self):
        return self.__convert_to_iterator(self.__collation_numbers_column_data)


# UI
class PathUI:
    def __init__(self, frame, settings):
        self.protocol_path = tk.StringVar()
        self.collation_path = tk.StringVar()
        self.frame = frame
        self.settings = settings
        Label(self.frame, text='Protokół').grid(column=0, row=0, sticky=W)
        Label(self.frame, text='Zestawienie').grid(column=0, row=1, sticky=W)
        Entry(self.frame, textvariable=self.protocol_path, width=60).grid(column=1, row=0, sticky=(W, E))
        Entry(self.frame, textvariable=self.collation_path, width=60).grid(column=1, row=1, sticky=(W, E))
        Button(self.frame, text='...', command=self.open_protocol).grid(column=2, row=0, sticky=E)
        Button(self.frame, text='...', command=self.open_collation).grid(column=2, row=1, sticky=E)
        Button(self.frame, text='Analiza', command=self.analyze).grid(column=0, row=2, columnspan=3, sticky=(W, E))

    def open_protocol(self):
        filetypes = (
            ('Excel', '*.xlsx'),
            ('All files', '*.*')
        )

        self.protocol_path.set(fd.askopenfilename(
            title='Open a file',
            initialdir='/',
            filetypes=filetypes))

    def open_collation(self):
        filetypes = (
            ('Excel', '*.xlsx'),
            ('All files', '*.*')
        )

        self.collation_path.set(fd.askopenfilename(
            title='Open a file',
            initialdir='/',
            filetypes=filetypes))

    def analyze(self):
        ready = True
        if self.protocol_path.get() == "":
            logger.log(logging.ERROR, "Brakująca ścieżka do protokołu")
            ready = False
        elif not self.protocol_path.get().endswith('.xlsx'):
            logger.log(logging.ERROR, "Podana sciezka dla protokolu nie jest plikiem excel (.xlsx)")
            ready = False

        if self.collation_path.get() == "":
            logger.log(logging.ERROR, "Brakująca ścieżka do zestawienia")
            ready = False
        elif not self.collation_path.get().endswith('.xlsx'):
            logger.log(logging.ERROR, "Podana sciezka dla zestawienia nie jest plikiem excel (.xlsx)")
            ready = False

        if ready:
            Analyzer(self.protocol_path.get(), self.collation_path.get(), self.settings)


class ConsoleUI:
    def __init__(self, frame):
        self.frame = frame
        self.scrolled_text = ScrolledText(frame, state='disabled', height=30, width=150)
        self.scrolled_text.grid(row=0, column=0, sticky=(N, S, W, E))
        self.log_queue = queue.Queue()
        self.queue_handler = QueueHandler(self.log_queue)
        logger.addHandler(self.queue_handler)
        self.frame.after(100, self.poll_log_queue)

    def clear_console(self):
        self.scrolled_text.delete(1.0, tk.END)

    def display(self, record):
        msg = self.queue_handler.format(record)
        self.scrolled_text.configure(state='normal')
        self.scrolled_text.insert(tk.END, msg + '\n', record.levelname)
        self.scrolled_text.configure(state='disabled')
        self.scrolled_text.yview(tk.END)

    def poll_log_queue(self):
        while True:
            try:
                record = self.log_queue.get(block=False)
            except queue.Empty:
                break
            else:
                self.display(record)
        self.frame.after(100, self.poll_log_queue)


class SettingsUI:
    def __init__(self, frame, settings: Settings):
        self.__frame = frame
        self.__settings = settings

        callback = self.__frame.register(self.callback_function)

        self.__protocol_names_column_data = StringVar()
        self.__collation_name_column_data = StringVar()
        self.__protocol_number_column_data = StringVar()
        self.__collation_numbers_column_data = StringVar()

        column_names_frame = ttk.LabelFrame(self.__frame, text="Kolumny")
        column_names_frame.columnconfigure(1, weight=1)
        column_names_frame.columnconfigure(2, weight=1)
        column_names_frame.grid(padx=10, pady=10, sticky=(W, E))

        Label(column_names_frame, text='Protokół').grid(column=1, row=0)
        Label(column_names_frame, text='Zestawienie').grid(column=2, row=0)
        Label(column_names_frame, text='Imię/Imiona').grid(column=0, row=1)
        Label(column_names_frame, text='Numer/y').grid(column=0, row=2)
        Entry(column_names_frame,
              textvariable=self.__protocol_names_column_data,
              validate="key",
              validatecommand=(callback, '%S')).grid(column=1, row=1, sticky=(E, W))
        Entry(column_names_frame,
              textvariable=self.__collation_name_column_data,
              validate="key",
              validatecommand=(callback, '%S')).grid(column=2, row=1, sticky=(E, W))
        Entry(column_names_frame,
              textvariable=self.__protocol_number_column_data,
              validate="key",
              validatecommand=(callback, '%S')).grid(column=1, row=2, sticky=(E, W))
        Entry(column_names_frame,
              textvariable=self.__collation_numbers_column_data,
              validate="key",
              validatecommand=(callback, '%S')).grid(column=2, row=2, sticky=(E, W))

    def callback_function(self, input_string: str):
        if input_string.isalpha():
            self.update_settings()
            return True
        else:
            return False

    def update_settings(self):
        self.__settings.update_settings(self.__protocol_number_column_data.get(),
                                      self.__protocol_names_column_data.get(),
                                      self.__collation_name_column_data.get(),
                                      self.__collation_numbers_column_data.get())


class App:

    def __init__(self, root):
        self.settings = Settings()

        self.root = root
        self.root.title('Anatool')
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        vertical_pane = ttk.PanedWindow(self.root, orient=VERTICAL)
        vertical_pane.grid(row=0, column=0, sticky="nsew")

        horizontal_pane = ttk.PanedWindow(vertical_pane, orient=HORIZONTAL)
        vertical_pane.add(horizontal_pane)

        left_pane = ttk.PanedWindow(horizontal_pane, orient=VERTICAL)
        horizontal_pane.add(left_pane)

        path_frame = ttk.LabelFrame(left_pane, text="Ścieżki")
        path_frame.columnconfigure(1, weight=1)
        left_pane.add(path_frame)

        settings_frame = ttk.LabelFrame(left_pane, text="Ustawienia")
        settings_frame.columnconfigure(0, weight=1)
        left_pane.add(settings_frame)

        console_frame = ttk.LabelFrame(horizontal_pane, text="Log")
        console_frame.columnconfigure(0, weight=1)
        console_frame.rowconfigure(0, weight=1)
        horizontal_pane.add(console_frame, weight=1)

        self.path = PathUI(path_frame, self.settings)
        self.console = ConsoleUI(console_frame)
        self.settings = SettingsUI(settings_frame, self.settings)


class Cmd:
    def __init__(self, argv):
        self.protocol = ''
        self.collation = ''
        self.argv = argv
        self.__settings = Settings()

        try:
            opts, args = getopt.getopt(self.argv, "hp:z:", ["protokol=", "zestawienie="])
        except getopt.GetoptError:
            print('main.py -p <sciezka_do_protokolu> -z <sciezka_do_zestawienia>')
            sys.exit(2)
        for opt, arg in opts:
            if opt == '-h':
                print('main.py -p <sciezka_do_protokolu> -z <sciezka_do_zestawienia>')
                sys.exit()
            elif opt in ("-p", "--protokol"):
                if arg.endswith('.xlsx'):
                    self.protocol = arg
                else:
                    print("Podana sciezka dla protokolu nie jest plikiem excel (.xlsx)")
                    sys.exit()
            elif opt in ("-z", "--zestawienie"):
                if arg.endswith('.xlsx'):
                    self.collation = arg
                else:
                    print("Podana sciezka dla protokolu nie jest plikiem excel (.xlsx)")
                    sys.exit()

        Analyzer(self.protocol, self.collation, self.__settings)


class Analyzer:
    def __init__(self, protocol_path: str, collation_path: str, settings: Settings):
        self.__protocol_path = protocol_path
        self.__collation_path = collation_path
        self.__settings = settings

        self.__sheet_protocol: openpyxl.workbook.workbook.Worksheet
        self.__sheet_collation: openpyxl.workbook.workbook.Worksheet

        self.collation = {}
        self.protocol = {}

        logger.log(logging.INFO, "----Analiza----")
        if not self.get_sheets():
            logger.log(logging.INFO, "----Analiza zakończona błędem----")
            return

        if not self.get_objects():
            logger.log(logging.INFO, "----Analiza zakończona błędem----")
            return

        if not self.analyze():
            logger.log(logging.INFO, "----Analiza zakończona błędem----")
            return

        logger.log(logging.INFO, "----Koniec analizy----")

    def get_sheets(self):
        try:
            wb_protocol = openpyxl.open(self.__protocol_path)
        except:
            logger.log(logging.ERROR, "Nie można otworzyć pliku protokołu")
            return False

        try:
            wb_collation = openpyxl.open(self.__collation_path)
        except:
            logger.log(logging.ERROR, "Nie można otworzyć pliku zestawienia")
            return False

        self.__sheet_protocol = wb_protocol[wb_protocol.sheetnames[0]]
        self.__sheet_collation = wb_collation[wb_collation.sheetnames[0]]

        return True

    def get_objects(self):

        # Protocol
        rows_from_sheet = self.__sheet_protocol.iter_rows()
        rows = iter(rows_from_sheet)
        for row in rows:
            number = str(row[self.__settings.protocol_number_column()].value)
            if number == 'None':
                continue

            str_names = row[self.__settings.protocol_names_column()].value
            # Split string to list
            split_names = str_names.split('\n')
            # Remove white spaces
            split_names = [n.strip() for n in split_names]
            # Remove empty strings (new lines)
            split_names = [n for n in split_names if n != '']

            if number in self.protocol:
                new_list = self.protocol[number].names
                new_list.extend(split_names)
                new_list = list(dict.fromkeys(new_list))
                self.protocol[number].names = new_list
            else:
                self.protocol[number] = Protocol(number, split_names, row[self.__settings.protocol_number_column()].row)

        # Collation
        current_row = 0
        rows_from_sheet = self.__sheet_collation.iter_rows()
        rows = iter(rows_from_sheet)
        for row in rows:
            current_row += 1
            try:
                name = row[self.__settings.collation_name_column()].value.rstrip()
                str_plot_numbers = row[self.__settings.collation_numbers_column()].value
                # Split string to list
                list_plot_numbers = str_plot_numbers.split(",")
                # Remove white spaces
                list_plot_numbers = [n.strip() for n in list_plot_numbers]
                # Remove empty strings (new lines)
                list_plot_numbers = [n for n in list_plot_numbers if n != '']
                if name in self.collation:
                    self.collation[name].add_positions(list_plot_numbers)
                else:
                    self.collation[name] = Collation(row[self.__settings.collation_name_column()].value,
                                                     list_plot_numbers, row[self.__settings.collation_name_column()].row)
            except:
                logger.log(logging.ERROR, "Błąd parsowania pliku zestawienia dla wiersza " + str(current_row))

        return True

    def analyze(self):
        for num, position in self.protocol.items():
            for name in position.names:
                not_found = -2

                if name in self.collation:
                    not_found = -1

                    if num in self.collation[name].numbers:
                        not_found = 0

                if not_found == -1:
                    msg = "Brakujaca pozycja " + position.number + " z protokołu dla nazwiska " + name + " w zestawieniu.; Linia w zestawieniu: " + str(
                        self.collation[name].row)
                    logger.log(logging.ERROR, msg)
                elif not_found == -2:
                    msg = "Brakujace nazwisko " + name + " w zestawieniu" + "; " \
                          + "Pozycja w protokole: " + position.number + "."
                    logger.log(logging.ERROR, msg)

        for name, collation in self.collation.items():
            for num in collation.numbers:
                if num not in self.protocol:
                    logger.log(logging.ERROR,
                               "Nieistniejąca pozycja " + str(
                                   num) + " protokołu w zestawieniu; Linia w protokole: " + str(collation.row))
                elif name not in self.protocol[num].names:
                    logger.log(logging.ERROR,
                               "Nazwisko " + name + " nie widnieje w protokole dla pozycji " + str(
                                   num) + " w zestawieniu; Linia w protokole: " + str(collation.row))


if __name__ == "__main__":
    if len(sys.argv) > 1:
        Cmd(sys.argv[1:])
    else:
        logging.basicConfig(level=logging.DEBUG)
        root = tk.Tk()
        debug_state = IntVar()
        app = App(root)
        app.root.mainloop()
